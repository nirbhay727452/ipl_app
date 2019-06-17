import openpyxl
import os
import sys
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE",'iplApp.settings')
django.setup()
from ipl.models import *



#
#
# wb_obj = openpyxl.load_workbook("matches.xlsx")
# sheet_obj = wb_obj.active
# row=sheet_obj.max_row
# col=sheet_obj.max_column
# for i in range(2,row+1):
#     if sheet_obj.cell(i,1).value==None:
#         continue
#     c=Matches(id=sheet_obj.cell(i,1).value,
#               season=sheet_obj.cell(i,2).value,
#               city=sheet_obj.cell(i,3).value,
#               date=sheet_obj.cell(i,4).value,
#               team1=sheet_obj.cell(i,5).value,
#               team2=sheet_obj.cell(i,6).value,
#               toss_winner=sheet_obj.cell(i,7).value,
#               toss_decision=sheet_obj.cell(i,8).value,
#               result=sheet_obj.cell(i,9).value,
#               dl_applied=sheet_obj.cell(i,10).value,
#               winner=sheet_obj.cell(i,11).value,
#               win_by_run=sheet_obj.cell(i,12).value,
#               win_by_wickets=sheet_obj.cell(i,13).value,
#               player_of_match=sheet_obj.cell(i,14).value,
#               venue=sheet_obj.cell(i,15).value,
#               umpire1=sheet_obj.cell(i,16).value,
#               umpire2=sheet_obj.cell(i,17).value,
#               umpire3=sheet_obj.cell(i,18).value
#               )
#     c.save()





wb_obj = openpyxl.load_workbook("deliveries.xlsx")
sheet_obj = wb_obj.active
row=sheet_obj.max_row
for i in range(2,row+1):
    if sheet_obj.cell(i,1).value==None:
        continue
    c=Deliveries(
            # match_id=sheet_obj.cell(i,1).value,
              match=Matches.objects.get(id=sheet_obj.cell(i,1).value),
              inning=sheet_obj.cell(i,2).value,
              batting_team=sheet_obj.cell(i,3).value,
              bowling_team=sheet_obj.cell(i,4).value,
              over=sheet_obj.cell(i,5).value,
              ball=sheet_obj.cell(i,6).value,
              batsman=sheet_obj.cell(i,7).value,
              non_striker=sheet_obj.cell(i,8).value,
              bowler=sheet_obj.cell(i,9).value,
              is_super_over=sheet_obj.cell(i,10).value,
              wide_runs=sheet_obj.cell(i,11).value,
              bye_runs=sheet_obj.cell(i,12).value,
              legbye_runs=sheet_obj.cell(i,13).value,
              noball_runs=sheet_obj.cell(i,14).value,
              penality_runs=sheet_obj.cell(i,15).value,
              batsman_runs=sheet_obj.cell(i,16).value,
              extra_runs=sheet_obj.cell(i,17).value,
              total_runs=sheet_obj.cell(i,18).value,
              player_dismissed=sheet_obj.cell(i,19).value,
              dismissal=sheet_obj.cell(i,20).value,
              fielder=sheet_obj.cell(i,21).value
              )
    c.save()
